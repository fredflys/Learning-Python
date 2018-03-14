import os
import sys
import shutil
from openpyxl import load_workbook


def get_user_input():
    rtc_num = input('Please input RTC number: ')
    jira_info = input('Please input Jira Info: ').split('-')
    iter_num = input('Please input Iteration number: ')
    release_date = input('Please input release date(yyyymmdd): ')
    cycle_num = input('Please input cycle number(1 or 2): ')
    try:
        int(rtc_num)
        int(iter_num)
        int(jira_info[1])
    except (ValueError, IndexError) as e:
        give_info("Invliad input. Behave yourself and don't mess with me.")
        return False
    jira_name = jira_info[0]
    jira_num = jira_info[1]
    return jira_name, jira_num, rtc_num, iter_num, release_date, cycle_num


def decide_src_file(jira_name, username):
    if jira_name.strip().upper() == 'CFP':
        src_path = r'c:\\Users\\' + username + '\\Sample.xlsx'
    elif jira_name.strip().upper() == 'CMR':
        src_path = r'c:\\Users\\' + username + '\\Sample For Report.xlsx'
    else:
        give_info('Invalid jira name. Show some manners, plz.')
        return False
    return src_path


def CreateTestFile(source_path, target_folder, jira_name, jira_num, rtc_num):
    if os.path.exists(source_path):
        target_path = target_folder + r'\(' + rtc_num + ')' + jira_name.upper() + '-' + jira_num + ' Test Cases and Test Results.xlsx'
        if os.path.exists(target_path):
            give_info('Test file already exists. So the copy will not be made.')
        else:
            shutil.copy(source_path, target_path)
    else:
        give_info('Original file or target file path does not exist.')
        return False
    return target_path


def revise_test_file(file_path, iter_num, rtc_num, release_date, cycle_num):
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        for col in ws.iter_cols(min_row=2, max_row=10, min_col=7, max_col=7):
            for cell in col:
                cell.value = r'CM_CH1_Agile\CURRENT RELEASES\CH_CH1_R' + release_date + r'\RTC_It' + iter_num + '\CH1_It' + iter_num + '_Cycle' + cycle_num + r'\Story ' + rtc_num
        wb.save(file_path)
    else:
        give_info('File path does not exist. No file will be revised.')
        return False
    return True


# 下面用装饰器修饰打印提示信息的函数，使其不会立刻退出，对用户更加友好
def pause(func):
    def wrapper(*args, **kwargs):
        result = func(*args, **kwargs)
        input('Please Enter Any Key to Quit...')
        return result

    return wrapper


@pause
def give_info(text):
    print(text)


def main():
    try:
        jira_name, jira_num, rtc_num, iter_num, release_date, cycle_num = get_user_input()
    except TypeError:
        give_info('Process denied due to invalid user inputs.')
        sys.exit()
    # 使用cmd命令，显示用户名，并读取结果，去掉后面的换行符，方便之后在桌面创建文件
    username = os.popen('echo %username%').read()[0:-1]
    target_folder = 'c:\\Users\\' + username + '\\Desktop'
    src_path = decide_src_file(jira_name, username)
    if not os.path.exists(src_path) or not os.path.exists(target_folder):
        give_info('Source test file not found.Process interrupted.')
        sys.exit()
    target_path = CreateTestFile(src_path, target_folder, jira_name, jira_num, rtc_num)
    if target_path:
        print('Test file created.')
        succeeds_revision = revise_test_file(target_path, iter_num, rtc_num, release_date, cycle_num)
    if succeeds_revision:
        give_info('Revision succeeds.')


main()