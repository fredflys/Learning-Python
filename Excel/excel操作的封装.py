import os
from openpyxl import Workbook
from openpyxl import load_workbook


class ExcelUtil(object):
    def __init__(self, excel_file_path):
        if os.path.exists(excel_file_path):
            wb = load_workbook(excel_file_path)
        else:
            wb = Workbook()
            wb.save(excel_file_path)
        self.excel_file_path = excel_file_path
        self.wb = wb

    def create_sheet(self, sheet_name):
        self.wb.create_sheet(sheet_name, 0)
        self.wb.save(self.excel_file_path)

    # 原模块中无法直接通过索引获得工作表，因此加一层抽象，间接封装一个索引得表的方法
    def get_sheet_by_index(self, index):
        try:
            return self.wb.get_sheet_by_name(self.wb.get_sheet_names()[index])
        except Exception as e_index:
            print("An error occurred when activating the sheet: ",e_index)

    def write_cell_value(self, index, column, row, value):
        try:
            current_ws = self.get_sheet_by_index(index)
            current_ws[column + row] = value
            self.wb.save(self.excel_file_path)
            print('Wrting succeeded.')
        except Exception as e_value:
            print("An error occurred when writing into a cell: ", e_value)


if __name__ == "__main__":
    eu = ExcelUtil("d:\demo.xlsx")
    print(eu.get_sheet_by_index(1))
    eu.write_cell_value(1, "A", "1", u"第二次添加的内容")
